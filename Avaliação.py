import openpyxl 
import datetime
from tabulate import tabulate
from termcolor import colored
from colorama import init
init(autoreset=True)


def informarData():
    mes = int(input("Insira o mês (1 a 12): "))
    while 1>mes or mes>12 :
        print("Mês invalido")
        mes = int(input("Insira o mês (1 a 12): "))
    ano = int(input("Insira o ano(23 ou 24): "))
    while 23>ano or ano>24:
        print("Ano invalido")
        ano = int(input("Insira o ano(em 2 digitos): "))
    return f"{mes}/{ano}:"

def acrescentarNaTabela(data, colunaAcrescentar, valor):
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    for linha in range(1, pagina.max_row +1):
        for coluna in "A":
            cellName = f"{coluna}{linha}"
            if pagina[cellName].value == data:
                cellData = f"{colunaAcrescentar}{linha}"
                pagina[cellData] = valor
    excel.save('Avaliação.xlsx') 
   
def acharData(data, colunaAchar):
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    for linha in range(1, pagina.max_row +1):
        for coluna in "A":
            cellName = f"{coluna}{linha}"
            if pagina[cellName].value == data:
                return f"{colunaAchar}{linha}" 

def validarDespesa(data, despesa):
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    cellSalario = acharData(data,"B")
    salarioDaData = pagina[cellSalario].value
    if despesa > salarioDaData:
        print("Não se pode adicionar uma despesa maior que o salario")
    else:
        acrescentarNaTabela(data, "C", despesa)
    
def informarSalario():
    data = informarData()
    salario = float(input(f"Informe o salario de {data} ")) 
    acrescentarNaTabela(data, "B", salario)       

def alterarSalario():
    print("informe a data que deseja alterar")
    data = informarData()
    salario = float(input("Informe o novo salario: "))
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    cellSalario = acharData(data,"C")
    despesaDaData = pagina[cellSalario].value
    if salario < despesaDaData:
        print("Não se pode adicionar uma despesa maior que o salario")
    else:
        acrescentarNaTabela(data, "B", salario)

def excluir():
    print("Informe a data que deseja excluir")
    data = informarData()
    acrescentarNaTabela(data, "B", "0")
    acrescentarNaTabela(data, "C", "0")
    
def listarSalarios():
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    tabela = []
    for linha in range(1, pagina.max_row + 1):
        cell = f"A{linha}"
        cell2 = f"B{linha}"
        tabela.append([pagina[cell].value, pagina[cell2].value])
    print(tabulate(tabela))

def informarDespesa():
    data = informarData()
    despesa = float(input(f"Informe a despesa de {data}"))
    validarDespesa(data, despesa)
  
def alterarDespesa():
    print("Informe a data que dejesa alterar")
    data = informarData()
    despesa = float(input(f"Informe a despesa de {data}"))
    validarDespesa(data, despesa)
    
def listarDespesas():
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    tabela = []
    for linha in range(1, pagina.max_row + 1):
        cell = f"A{linha}"
        cell2 = f"C{linha}"
        tabela.append([pagina[cell].value, pagina[cell2].value])
    print(tabulate(tabela))
    
def calcularInvestimento():
    excel = openpyxl.load_workbook('Avaliação.xlsx')    
    pagina = excel.active
    for linha in range(2, pagina.max_row + 1 ):
        for coluna in "BC":
            cell = f'D{linha}'
            salario = pagina[f'B{linha}'].value
            despesa = pagina[f'C{linha}'].value
            salarioFloat = float(salario)
            despesaFloat = float(despesa)
            investimento = salarioFloat - despesaFloat
            pagina[cell] = investimento
    excel.save("Avaliação.xlsx")

            

def sair():
    exit()
    
    
def menu_de_opcoes():
   print()

   print(colored("Digite a opção desejada: ", "white", "on_yellow"))
   print(colored("1 - Informar salario", "yellow"))
   print(colored("2 - Alterar salario", "yellow"))
   print(colored("3 - Excluir salario", "yellow"))
   print(colored("4 - Listar salarios", "yellow"))
   print(colored("5 - Informar despesa", "yellow"))
   print(colored("6 - Alterar despesa", "yellow"))
   print(colored("7 - Remover despesa", "yellow"))
   print(colored("8 - Listar despesas", "yellow"))
   print(colored("9 - Calcular Investimento", "yellow"))
   print(colored("0 - Sair", "yellow"))

   print()

   try:
       opcao_selecionada = int(input(colored("joao-serasa: ~$ ", "light_green")))
   except:
       print(colored("Erro: Opção informada inválida", "red"))
       return menu_de_opcoes()
   if 0 <= opcao_selecionada < 10:
       return opcao_selecionada
   else:
       print(colored("Erro: Opção informada inválida", "red"))
       return menu_de_opcoes()

while True:
       [sair, informarSalario, alterarSalario, excluir,
        listarSalarios, informarDespesa, alterarDespesa, excluir,
        listarDespesas, calcularInvestimento][menu_de_opcoes()]()
   
